import os
import json
import re
import io
import pandas as pd
import pdfplumber
import uuid 
from datetime import datetime
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from flask import Blueprint, render_template, request, flash, redirect, url_for, send_file, current_app
from flask_login import login_required, current_user
from models import db, Mapel, Ujian, JawabanSiswa, User
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from xhtml2pdf import pisa

bp = Blueprint('guru', __name__)

# ==================== HELPER: PARSE PDF LINES ====================
def parse_pdf_lines(lines):
    pg_list = []
    essay_list = []
    current_soal = None
    last_state = 'soal'

    # Regex Patterns
    pola_nomor = re.compile(r'^(\d+)\.\s+(.*)')
    pola_opsi = re.compile(r'^([A-E])\.\s+(.*)')
    pola_kunci = re.compile(r'\(Jawaban:\s*([A-E])\)', re.IGNORECASE)
    pola_bobot = re.compile(r'\(Poin:\s*(\d+)\)', re.IGNORECASE)

    for line in lines:
        temu_kunci = pola_kunci.search(line)
        temu_bobot = pola_bobot.search(line)

        line_clean = line
        found_key_val = None
        found_bobot_val = None

        if temu_kunci:
            found_key_val = temu_kunci.group(1).upper()
            line_clean = pola_kunci.sub('', line_clean).strip()

        if temu_bobot:
            found_bobot_val = int(temu_bobot.group(1))
            line_clean = pola_bobot.sub('', line_clean).strip()

        if not line_clean and not found_key_val and not found_bobot_val:
            continue

        match_nomor = pola_nomor.match(line_clean)
        match_opsi = pola_opsi.match(line_clean)

        if match_nomor:
            if current_soal:
                if current_soal['tipe'] == 'pg':
                    pg_list.append(current_soal['data'])
                elif current_soal['tipe'] == 'essay':
                    essay_list.append(current_soal['data'])

            isi_soal = match_nomor.group(2).strip()
            
            # GENERATE ID UNIK UNTUK SOAL BARU DARI PDF
            new_id = str(uuid.uuid4())

            if found_key_val:
                current_soal = {
                    'tipe': 'pg',
                    'data': {
                        'id': new_id,
                        'soal': isi_soal,
                        'a': '', 'b': '', 'c': '', 'd': '', 'e': '',
                        'kunci': found_key_val,
                        'gambar': ''
                    }
                }
            elif found_bobot_val:
                current_soal = {
                    'tipe': 'essay',
                    'data': {'id': new_id, 'soal': isi_soal, 'bobot': found_bobot_val, 'gambar': ''}
                }
            else:
                current_soal = {
                    'tipe': 'pg',
                    'data': {
                        'id': new_id,
                        'soal': isi_soal,
                        'a': '', 'b': '', 'c': '', 'd': '', 'e': '',
                        'kunci': 'A',
                        'gambar': ''
                    }
                }
            last_state = 'soal'

        elif current_soal and (found_key_val or found_bobot_val):
            if found_key_val and current_soal['tipe'] == 'pg':
                current_soal['data']['kunci'] = found_key_val

            if found_bobot_val:
                if current_soal['tipe'] == 'essay':
                    current_soal['data']['bobot'] = found_bobot_val
                elif current_soal['tipe'] == 'pg':
                    current_soal['tipe'] = 'essay'
                    current_soal['data'] = {
                        'id': current_soal['data']['id'],
                        'soal': current_soal['data']['soal'],
                        'bobot': found_bobot_val,
                        'gambar': ''
                    }

            if line_clean:
                if last_state == 'soal':
                    current_soal['data']['soal'] += " " + line_clean
                elif last_state in ['a', 'b', 'c', 'd', 'e'] and current_soal['tipe'] == 'pg':
                    current_soal['data'][last_state] += " " + line_clean

        elif current_soal and current_soal['tipe'] == 'pg' and match_opsi:
            opt_label = match_opsi.group(1).lower()
            opt_text = match_opsi.group(2).strip()
            current_soal['data'][opt_label] = opt_text
            last_state = opt_label

        elif current_soal and line_clean:
            if last_state == 'soal':
                current_soal['data']['soal'] += " " + line_clean
            elif last_state in ['a', 'b', 'c', 'd', 'e'] and current_soal['tipe'] == 'pg':
                current_soal['data'][last_state] += " " + line_clean

    if current_soal:
        if current_soal['tipe'] == 'pg':
            pg_list.append(current_soal['data'])
        elif current_soal['tipe'] == 'essay':
            essay_list.append(current_soal['data'])

    return pg_list, essay_list


# ==================== DASHBOARD GURU ====================
@bp.route('/dashboard')
@login_required
def dashboard():
    if current_user.role not in ['guru', 'admin']:
        return redirect('/')

    if current_user.role == 'admin':
        mapel = Mapel.query.all()
        ujian = Ujian.query.order_by(Ujian.waktu_mulai.desc()).all()
    else:
        mapel = Mapel.query.filter_by(guru_id=current_user.id).all()
        ujian = Ujian.query.join(Mapel).filter(Mapel.guru_id == current_user.id).order_by(
            Ujian.waktu_mulai.desc()).all()

    return render_template('guru/dashboard.html', mapel=mapel, ujian=ujian, datetime=datetime)


# ==================== UPLOAD SOAL / BUAT UJIAN BARU ====================
@bp.route('/upload_soal/<int:mapel_id>', methods=['GET', 'POST'])
@login_required
def upload_soal(mapel_id):
    if current_user.role not in ['guru', 'admin']:
        return redirect('/')

    mapel = Mapel.query.get_or_404(mapel_id)

    if current_user.role != 'admin' and mapel.guru_id != current_user.id:
        flash('Anda tidak berhak mengunggah soal untuk mapel ini!', 'danger')
        return redirect('/guru/dashboard')

    if request.method == 'POST':
        judul = request.form['judul'].strip()
        waktu_mulai = request.form['waktu_mulai']
        waktu_selesai = request.form['waktu_selesai']
        durasi_input = int(request.form['durasi'])
        file = request.files.get('file_pdf')

        if not judul or not waktu_mulai or not waktu_selesai:
            flash('Judul, Waktu Mulai, dan Waktu Selesai wajib diisi!', 'danger')
            return redirect(request.url)

        try:
            mulai = datetime.strptime(waktu_mulai, '%Y-%m-%dT%H:%M')
            selesai = datetime.strptime(waktu_selesai, '%Y-%m-%dT%H:%M')
        except ValueError:
            flash('Format waktu salah!', 'danger')
            return redirect(request.url)

        pg_list = []
        essay_list = []

        if file and file.filename != '':
            if not file.filename.lower().endswith('.pdf'):
                flash('File harus berformat PDF (.pdf)', 'danger')
                return redirect(request.url)

            try:
                full_text = ""
                with pdfplumber.open(file) as pdf:
                    for page in pdf.pages:
                        extracted = page.extract_text()
                        if extracted:
                            full_text += extracted + "\n"

                lines = [line.strip() for line in full_text.split('\n') if line.strip()]
                pg_list, essay_list = parse_pdf_lines(lines)

                if not pg_list and not essay_list:
                    flash('Gagal membaca soal! Pastikan format PDF sesuai.', 'warning')
                    return redirect(request.url)

            except Exception as e:
                flash(f'Terjadi kesalahan sistem saat membaca PDF: {str(e)}', 'danger')
                return redirect(request.url)

        ujian = Ujian(
            mapel_id=mapel_id,
            judul=judul,
            waktu_mulai=mulai,
            waktu_selesai=selesai,
            durasi_menit=durasi_input,
            soal_pg=json.dumps(pg_list, ensure_ascii=False),
            soal_essay=json.dumps(essay_list, ensure_ascii=False)
        )
        db.session.add(ujian)
        db.session.commit()

        if file and file.filename != '':
            flash(f'Berhasil! {len(pg_list)} Soal PG dan {len(essay_list)} Soal Essay tersimpan.', 'success')
            return redirect('/guru/dashboard')
        else:
            flash('Kerangka ujian berhasil dibuat! Silakan tambahkan soal secara manual.', 'success')
            return redirect(url_for('guru.edit_ujian', ujian_id=ujian.id))

    return render_template('guru/upload_soal.html', mapel=mapel)


# ==================== EDIT UJIAN (FIX: ID UNIK + RECALCULATE) ====================
@bp.route('/edit_ujian/<int:ujian_id>', methods=['GET', 'POST'])
@login_required
def edit_ujian(ujian_id):
    if current_user.role not in ['guru', 'admin']:
        return redirect('/')

    ujian = Ujian.query.get_or_404(ujian_id)

    if current_user.role != 'admin' and ujian.mapel.guru_id != current_user.id:
        flash('Akses ditolak!', 'danger')
        return redirect('/guru/dashboard')

    try:
        pg_existing = json.loads(ujian.soal_pg) if ujian.soal_pg else []
        essay_existing = json.loads(ujian.soal_essay) if ujian.soal_essay else []

    except:
        pg_existing = []
        essay_existing = []

    if request.method == 'POST':
        ujian.judul = request.form['judul'].strip()
        ujian.durasi_menit = int(request.form['durasi'])
        try:
            ujian.waktu_mulai = datetime.strptime(request.form['waktu_mulai'], '%Y-%m-%dT%H:%M')
            ujian.waktu_selesai = datetime.strptime(request.form['waktu_selesai'], '%Y-%m-%dT%H:%M')
        except ValueError:
            flash('Format waktu salah!', 'danger')
            return redirect(request.url)

        file = request.files.get('file_pdf')
        
        # Variabel untuk menampung soal baru (agar bisa dipakai hitung ulang nilai)
        final_pg_list = []
        final_essay_list = []

        # --- A. EDIT DENGAN UPLOAD PDF BARU ---
        if file and file.filename != '':
            if not file.filename.lower().endswith('.pdf'):
                flash('File harus format PDF!', 'danger')
                return redirect(request.url)

            try:
                full_text = ""
                with pdfplumber.open(file) as pdf:
                    for page in pdf.pages:
                        extracted = page.extract_text()
                        if extracted:
                            full_text += extracted + "\n"

                lines = [line.strip() for line in full_text.split('\n') if line.strip()]
                new_pg, new_essay = parse_pdf_lines(lines)

                if not new_pg and not new_essay:
                    flash('Gagal membaca PDF! Format tidak sesuai.', 'warning')
                    return redirect(request.url)

                ujian.soal_pg = json.dumps(new_pg, ensure_ascii=False)
                ujian.soal_essay = json.dumps(new_essay, ensure_ascii=False)
                
                final_pg_list = new_pg
                final_essay_list = new_essay
                
                flash('Soal diperbarui dari PDF & Nilai siswa sedang dihitung ulang...', 'success')

            except Exception as e:
                flash(f'Error membaca PDF: {str(e)}', 'danger')
                return redirect(request.url)

        # --- B. EDIT MANUAL (DENGAN ID PERMANEN) ---
        else:
            upload_folder = os.path.join(current_app.root_path, 'static', 'uploads', 'soal')
            if not os.path.exists(upload_folder):
                os.makedirs(upload_folder)

            # === 1. PILIHAN GANDA ===
            pg_ids = request.form.getlist('pg_id[]')
            soal_pg_texts = request.form.getlist('pg_soal[]')
            kunci_pg = request.form.getlist('pg_kunci[]')
            
            opsi_texts = {k: request.form.getlist(f'pg_{k}[]') for k in ['a','b','c','d','e']}

            manual_pg_list = []

            for i, uid in enumerate(pg_ids):
                status_img = request.form.get(f'pg_img_status_{uid}')
                file_img = request.files.get(f'pg_img_{uid}')
                old_img = request.form.get(f'pg_old_img_{uid}')

                gambar_final = ""
                if status_img == '1':
                    if file_img and file_img.filename:
                        filename = secure_filename(f"PG_{ujian_id}_{uid}_{int(datetime.now().timestamp())}_{file_img.filename}")
                        file_img.save(os.path.join(upload_folder, filename))
                        gambar_final = filename
                    else:
                        gambar_final = old_img
                
                item_soal = {
                    'id': uid, # SIMPAN ID (Bisa UUID baru atau Index "0" dari data lama)
                    'soal': soal_pg_texts[i],
                    'kunci': kunci_pg[i],
                    'gambar': gambar_final
                }

                for kode in ['a', 'b', 'c', 'd', 'e']:
                    item_soal[kode] = opsi_texts[kode][i]
                    
                    opt_status = request.form.get(f'pg_img_status_{kode}_{uid}')
                    opt_file = request.files.get(f'pg_img_{kode}_{uid}')
                    opt_old = request.form.get(f'pg_old_img_{kode}_{uid}')
                    
                    opt_img_final = ""
                    if opt_status == '1':
                        if opt_file and opt_file.filename:
                            fname = secure_filename(f"OPT_{kode}_{ujian_id}_{uid}_{int(datetime.now().timestamp())}_{opt_file.filename}")
                            opt_file.save(os.path.join(upload_folder, fname))
                            opt_img_final = fname
                        else:
                            opt_img_final = opt_old
                    
                    item_soal[f"{kode}_gambar"] = opt_img_final

                manual_pg_list.append(item_soal)

            # === 2. ESSAY ===
            essay_ids = request.form.getlist('essay_id[]')
            soal_essay_texts = request.form.getlist('essay_soal[]')
            bobot_essay = request.form.getlist('essay_bobot[]')

            manual_essay_list = []

            for i, uid in enumerate(essay_ids):
                status_img = request.form.get(f'essay_img_status_{uid}')
                file_img = request.files.get(f'essay_img_{uid}')
                old_img = request.form.get(f'essay_old_img_{uid}')

                gambar_final = ""
                if status_img == '1':
                    if file_img and file_img.filename:
                        filename = secure_filename(f"ES_{ujian_id}_{uid}_{int(datetime.now().timestamp())}_{file_img.filename}")
                        file_img.save(os.path.join(upload_folder, filename))
                        gambar_final = filename
                    else:
                        gambar_final = old_img
                
                manual_essay_list.append({
                    'id': uid, # SIMPAN ID
                    'soal': soal_essay_texts[i],
                    'bobot': int(bobot_essay[i]) if bobot_essay[i] else 0,
                    'gambar': gambar_final
                })

            ujian.soal_pg = json.dumps(manual_pg_list, ensure_ascii=False)
            ujian.soal_essay = json.dumps(manual_essay_list, ensure_ascii=False)
            
            final_pg_list = manual_pg_list
            final_essay_list = manual_essay_list
            
            flash('Perubahan tersimpan!', 'success')

        # ==================== HITUNG ULANG NILAI OTOMATIS ====================
        # Hitung Max Score PG Baru
        total_bobot_essay = sum(int(e.get('bobot', 0)) for e in final_essay_list)
        max_pg = 100 - total_bobot_essay
        if max_pg < 0: max_pg = 0
        
        all_jawaban = JawabanSiswa.query.filter_by(ujian_id=ujian_id).all()
        count_updated = 0
        
        for jwb in all_jawaban:
            user_pg_answers = json.loads(jwb.jawaban_pg) if jwb.jawaban_pg else {}
            jml_benar = 0
            
            for idx, soal in enumerate(final_pg_list):
                # 1. Coba ambil jawaban pakai ID (Prioritas Utama)
                soal_id = soal.get('id')
                jawaban_user = user_pg_answers.get(soal_id)
                
                # 2. Fallback: Jika ID tidak ketemu (Data lama), coba pakai Index Loop
                if jawaban_user is None:
                    # Cek apakah user menyimpan dengan key angka string "0", "1"
                    jawaban_user = user_pg_answers.get(str(idx))

                if jawaban_user == soal.get('kunci'):
                    jml_benar += 1
            
            if final_pg_list:
                nilai_pg_baru = (jml_benar / len(final_pg_list)) * max_pg
            else:
                nilai_pg_baru = 0
            
            jwb.nilai_pg = round(nilai_pg_baru, 2)
            jwb.total_nilai = jwb.nilai_pg + jwb.nilai_essay
            count_updated += 1
            
        if count_updated > 0:
            flash(f'Sukses! Nilai {count_updated} siswa telah dihitung ulang otomatis.', 'info')

        db.session.commit()
        return redirect('/guru/dashboard')

    return render_template('guru/edit_ujian.html',
                           ujian=ujian,
                           pg_list=pg_existing,
                           essay_list=essay_existing)


# ==================== HAPUS UJIAN ====================
@bp.route('/hapus_ujian/<int:ujian_id>', methods=['POST'])
@login_required
def hapus_ujian(ujian_id):
    if current_user.role not in ['guru', 'admin']:
        return redirect('/')

    ujian = Ujian.query.get_or_404(ujian_id)

    if current_user.role != 'admin' and ujian.mapel.guru_id != current_user.id:
        flash('Akses ditolak!', 'danger')
        return redirect('/guru/dashboard')

    try:
        JawabanSiswa.query.filter_by(ujian_id=ujian_id).delete()
        db.session.delete(ujian)
        db.session.commit()
        flash('Ujian berhasil dihapus beserta data jawabannya!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal menghapus ujian: {str(e)}', 'danger')

    return redirect('/guru/dashboard')


# ==================== PREVIEW UJIAN ====================
@bp.route('/preview/<int:ujian_id>')
@login_required
def preview(ujian_id):
    if current_user.role not in ['guru', 'admin']:
        return redirect('/')

    ujian = Ujian.query.get_or_404(ujian_id)

    if current_user.role != 'admin' and ujian.mapel.guru_id != current_user.id:
        flash('Akses ditolak!', 'danger')
        return redirect('/guru/dashboard')

    pg = json.loads(ujian.soal_pg) if ujian.soal_pg else []
    essay = json.loads(ujian.soal_essay) if ujian.soal_essay else []
    return render_template('guru/preview_ujian.html', ujian=ujian, pg=pg, essay=essay)


# ==================== KOREKSI (ROBUST ID LOOKUP) ====================
@bp.route('/koreksi/<int:jawaban_id>', methods=['GET', 'POST'])
@login_required
def koreksi(jawaban_id):
    if current_user.role not in ['guru', 'admin']:
        return redirect('/')

    jawaban_siswa = JawabanSiswa.query.get_or_404(jawaban_id)
    ujian = jawaban_siswa.ujian

    if current_user.role != 'admin' and ujian.mapel.guru_id != current_user.id:
        flash('Akses ditolak!', 'danger')
        return redirect('/guru/dashboard')

    soal_pg = json.loads(ujian.soal_pg) if ujian.soal_pg else []
    soal_essay = json.loads(ujian.soal_essay) if ujian.soal_essay else []
    
    jawab_pg = json.loads(jawaban_siswa.jawaban_pg) if jawaban_siswa.jawaban_pg else {}
    jawab_essay = json.loads(jawaban_siswa.jawaban_essay) if jawaban_siswa.jawaban_essay else {}

    # HITUNG JUMLAH BENAR PG (ROBUST)
    jml_benar = 0
    for idx, s in enumerate(soal_pg):
        # Cek by ID -> Fallback by Index
        ans = jawab_pg.get(s.get('id'))
        if ans is None: ans = jawab_pg.get(str(idx))
            
        if ans == s.get('kunci'):
            jml_benar += 1

    if request.method == 'POST':
        total_skor_essay = 0
        for i, soal in enumerate(soal_essay):
            try:
                skor = float(request.form.get(f'nilai_{i}', 0))
                if skor < 0: skor = 0
            except:
                skor = 0
            total_skor_essay += skor

        # Update PG Score juga (agar sinkron jika kunci berubah)
        total_bobot_essay_soal = sum(int(s.get('bobot', 0)) for s in soal_essay)
        max_pg_score = 100 - total_bobot_essay_soal
        if max_pg_score < 0: max_pg_score = 0
        
        nilai_pg_baru = 0
        if soal_pg:
            nilai_pg_baru = (jml_benar / len(soal_pg)) * max_pg_score

        jawaban_siswa.nilai_pg = round(nilai_pg_baru, 2)
        jawaban_siswa.nilai_essay = total_skor_essay
        jawaban_siswa.total_nilai = jawaban_siswa.nilai_pg + total_skor_essay
        db.session.commit()
        
        flash(f'Nilai berhasil disimpan! (PG: {jawaban_siswa.nilai_pg}, Essay: {total_skor_essay})', 'success')
        return redirect(url_for('guru.lihat_nilai', ujian_id=ujian.id))

    # MAPPING JAWABAN KE LIST (AGAR TEMPLATE MUDAH)
    mapped_jawab_pg = {}
    for i, s in enumerate(soal_pg):
        ans = jawab_pg.get(s.get('id'))
        if ans is None: ans = jawab_pg.get(str(i))
        mapped_jawab_pg[str(i)] = ans
        
    mapped_jawab_essay = {}
    for i, s in enumerate(soal_essay):
        ans = jawab_essay.get(s.get('id'))
        if ans is None: ans = jawab_essay.get(str(i))
        mapped_jawab_essay[str(i)] = ans

    return render_template('guru/koreksi.html',
                           jawaban=jawaban_siswa,
                           soal_pg=soal_pg,
                           soal_essay=soal_essay,
                           jawab_pg=mapped_jawab_pg,
                           jawab_essay=mapped_jawab_essay,
                           jml_benar_pg=jml_benar, 
                           total_soal_pg=len(soal_pg)) 


# ==================== LIHAT NILAI (ROBUST CALCULATION) ====================
@bp.route('/lihat_nilai/<int:ujian_id>', methods=['GET', 'POST'])
@login_required
def lihat_nilai(ujian_id):
    if current_user.role not in ['guru', 'admin']:
        return redirect('/')

    ujian = Ujian.query.get_or_404(ujian_id)

    if current_user.role != 'admin' and ujian.mapel.guru_id != current_user.id:
        flash('Anda tidak memiliki akses ke data ini!', 'danger')
        return redirect('/guru/dashboard')

    data_nilai = JawabanSiswa.query.filter_by(ujian_id=ujian_id).all()

    soal_pg_raw = json.loads(ujian.soal_pg) if ujian.soal_pg else []
    total_pg = len(soal_pg_raw)

    for n in data_nilai:
        jw_pg = json.loads(n.jawaban_pg) if n.jawaban_pg else {}
        benar = 0
        for idx, s in enumerate(soal_pg_raw):
            # Cek by ID -> Fallback by Index
            ans = jw_pg.get(s.get('id'))
            if ans is None: ans = jw_pg.get(str(idx))
            
            if ans == s.get('kunci'):
                benar += 1

        n.jml_benar_pg = benar
        n.total_soal_pg = total_pg

    data_nilai.sort(
        key=lambda x: (x.siswa.kelas.nama_kelas if x.siswa and x.siswa.kelas else "", x.siswa.nama if x.siswa else ""))

    if request.method == 'POST' and 'download_excel' in request.form:
        if not data_nilai:
            flash('Belum ada siswa yang mengerjakan.', 'warning')
            return redirect(request.url)

        try:
            list_data = []
            for j in data_nilai:
                list_data.append({
                    'No': 0,
                    'NIS': j.siswa.username if j.siswa else '-',
                    'Nama Siswa': j.siswa.nama if j.siswa else '-',
                    'Kelas': j.siswa.kelas.nama_kelas if j.siswa and j.siswa.kelas else '-',
                    'Jml Benar PG': getattr(j, 'jml_benar_pg', 0),  
                    'Nilai PG': j.nilai_pg,
                    'Nilai Essay': j.nilai_essay,
                    'Total Nilai': j.total_nilai,
                    'Waktu Submit': j.waktu_submit.strftime('%Y-%m-%d %H:%M') if j.waktu_submit else '-'
                })

            for idx, item in enumerate(list_data, 1):
                item['No'] = idx

            df = pd.DataFrame(list_data)
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Nilai Ujian', startrow=5)
                workbook = writer.book
                worksheet = writer.sheets['Nilai Ujian']

                font_std = Font(name='Times New Roman', size=12)
                font_bold = Font(name='Times New Roman', size=12, bold=True)
                font_title = Font(name='Times New Roman', size=14, bold=True)
                border_thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                border_bottom_thick = Border(bottom=Side(style='medium'))

                worksheet.merge_cells('A1:B3')
                logo_path = os.path.join(current_app.root_path, 'static', 'img', 'logo_sekolah.png')
                if os.path.exists(logo_path):
                    img = ExcelImage(logo_path)
                    img.height = 70
                    img.width = 70
                    worksheet.add_image(img, 'A1')
                    worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

                worksheet.merge_cells('C1:I1')  
                cell_sekolah = worksheet['C1']
                cell_sekolah.value = "SMA ISLAM PLUS BAITUSSALAM"
                cell_sekolah.font = font_title
                cell_sekolah.alignment = Alignment(horizontal="center", vertical="bottom")

                worksheet.merge_cells('C2:I2')  
                cell_judul = worksheet['C2']
                cell_judul.value = f"LAPORAN HASIL UJIAN: {ujian.judul.upper()}"
                cell_judul.font = font_bold
                cell_judul.alignment = Alignment(horizontal="center", vertical="center")

                worksheet.merge_cells('C3:I3')  
                cell_info = worksheet['C3']
                cell_info.value = f"Mapel: {ujian.mapel.nama} | Tanggal Cetak: {datetime.now().strftime('%d %B %Y')}"
                cell_info.font = Font(name='Times New Roman', size=11, italic=True)
                cell_info.alignment = Alignment(horizontal="center", vertical="top")

                for col in range(1, 10):
                    cell = worksheet.cell(row=3, column=col)
                    cell.border = border_bottom_thick

                header_row = 6
                for i, col in enumerate(df.columns):
                    col_idx = i + 1
                    col_letter = get_column_letter(col_idx)
                    max_len = len(str(col))
                    for cell in worksheet[col_letter]:
                        if cell.row > header_row:
                            if cell.value:
                                max_len = max(max_len, len(str(cell.value)))
                            cell.font = font_std
                            cell.border = border_thin
                            if col in ['No', 'Kelas', 'Jml Benar PG', 'Nilai PG', 'Nilai Essay', 'Total Nilai']:
                                cell.alignment = Alignment(horizontal="center")

                    worksheet.column_dimensions[col_letter].width = max_len + 4
                    cell_header = worksheet.cell(row=header_row, column=col_idx)
                    cell_header.value = col
                    cell_header.font = font_bold
                    cell_header.alignment = Alignment(horizontal="center", vertical="center")
                    cell_header.border = border_thin
                    cell_header.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

            output.seek(0)
            safe_judul = secure_filename(ujian.judul)
            if not safe_judul: safe_judul = f"Ujian_{ujian.id}"
            filename = f"Rekap_{safe_judul}.xlsx"

            return send_file(output, as_attachment=True, download_name=filename,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        except Exception as e:
            flash(f"Gagal membuat Excel: {str(e)}", "danger")
            return redirect(request.url)

    return render_template('guru/lihat_nilai.html', ujian=ujian, data_nilai=data_nilai)


# ==================== HTMX REFRESH ====================
@bp.route('/refresh_tabel_nilai/<int:ujian_id>')
@login_required
def refresh_tabel_nilai(ujian_id):
    if current_user.role not in ['guru', 'admin']:
        return ('', 403)

    ujian = Ujian.query.get_or_404(ujian_id)
    if current_user.role != 'admin' and ujian.mapel.guru_id != current_user.id:
        return ('', 403)

    data_nilai = JawabanSiswa.query.filter_by(ujian_id=ujian_id).all()

    soal_pg_raw = json.loads(ujian.soal_pg) if ujian.soal_pg else []
    total_pg = len(soal_pg_raw)

    for n in data_nilai:
        jw_pg = json.loads(n.jawaban_pg) if n.jawaban_pg else {}
        benar = 0
        for idx, s in enumerate(soal_pg_raw):
            # Cek by ID -> Fallback by Index
            ans = jw_pg.get(s.get('id'))
            if ans is None: ans = jw_pg.get(str(idx))
            
            if ans == s.get('kunci'):
                benar += 1
        n.jml_benar_pg = benar
        n.total_soal_pg = total_pg

    data_nilai.sort(
        key=lambda x: (x.siswa.kelas.nama_kelas if x.siswa and x.siswa.kelas else "", x.siswa.nama if x.siswa else ""))
    return render_template('guru/partials/tabel_nilai_body.html', data_nilai=data_nilai)


# ==================== RESET PESERTA ====================
@bp.route('/reset_peserta/<int:jawaban_id>', methods=['POST'])
@login_required
def reset_peserta(jawaban_id):
    if current_user.role not in ['guru', 'admin']:
        return ('', 403)

    jawaban = JawabanSiswa.query.get_or_404(jawaban_id)
    if current_user.role != 'admin' and jawaban.ujian.mapel.guru_id != current_user.id:
        return ('', 403)

    db.session.delete(jawaban)
    db.session.commit()
    return ('', 204)

# ==================== DOWNLOAD PDF HASIL (FINAL) ====================
@bp.route('/download_hasil_pdf/<int:jawaban_id>')
@login_required
def download_hasil_pdf(jawaban_id):
    if current_user.role not in ['guru', 'admin']:
        return redirect('/')

    try:
        jawaban = JawabanSiswa.query.get_or_404(jawaban_id)
        ujian = jawaban.ujian
        siswa = jawaban.siswa

        if current_user.role != 'admin' and ujian.mapel.guru_id != current_user.id:
            flash('Anda tidak memiliki akses ke ujian ini.', 'danger')
            return redirect('/guru/dashboard')

        try: soal_pg = json.loads(ujian.soal_pg) if ujian.soal_pg else []
        except: soal_pg = []
        
        try: soal_essay = json.loads(ujian.soal_essay) if ujian.soal_essay else []
        except: soal_essay = []

        try: jawab_pg = json.loads(jawaban.jawaban_pg) if jawaban.jawaban_pg else {}
        except: jawab_pg = {}

        try: jawab_essay = json.loads(jawaban.jawaban_essay) if jawaban.jawaban_essay else {}
        except: jawab_essay = {}

        jml_benar_pg = 0
        for idx, s in enumerate(soal_pg):
            # Cek by ID -> Fallback by Index
            ans = jawab_pg.get(s.get('id'))
            if ans is None: ans = jawab_pg.get(str(idx))
            
            if ans == s.get('kunci'):
                jml_benar_pg += 1

        image_folder = os.path.join(current_app.root_path, 'static', 'uploads', 'soal')
        image_folder = image_folder.replace('\\', '/')

        html_content = render_template(
            'guru/pdf_hasil_siswa.html',
            siswa=siswa, ujian=ujian, jawaban=jawaban,
            soal_pg=soal_pg, soal_essay=soal_essay,
            jawab_pg=jawab_pg, jawab_essay=jawab_essay,
            jml_benar_pg=jml_benar_pg, total_soal_pg=len(soal_pg),
            image_folder=image_folder
        )

        pdf_output = io.BytesIO()
        pisa_status = pisa.CreatePDF(src=html_content, dest=pdf_output)

        if pisa_status.err:
            flash(f'Gagal membuat PDF: {pisa_status.err}', 'danger')
            return redirect(url_for('guru.lihat_nilai', ujian_id=ujian.id))

        pdf_output.seek(0)
        filename = secure_filename(f"Hasil_{siswa.nama}_{ujian.judul}.pdf")
        
        return send_file(
            pdf_output, 
            as_attachment=True, 
            download_name=filename, 
            mimetype='application/pdf'
        )

    except Exception as e:
        print(f"ERROR PDF: {e}") 
        flash("Terjadi kesalahan sistem saat membuat PDF.", "danger")
        return redirect('/guru/dashboard')
    
# ==================== GANTI PASSWORD ====================
@bp.route('/ganti_password', methods=['GET', 'POST'])
@login_required
def ganti_password():
    if current_user.role not in ['guru', 'admin']:
        return redirect('/')

    if request.method == 'POST':
        old_pass = request.form['old_pass']
        new_pass = request.form['new_pass']
        confirm_pass = request.form['confirm_pass']
        if not check_password_hash(current_user.password, old_pass):
            flash('Password lama salah!', 'danger')
        elif new_pass != confirm_pass:
            flash('Konfirmasi password baru tidak cocok!', 'warning')
        elif len(new_pass) < 6:
            flash('Password baru minimal 6 karakter!', 'warning')
        else:
            current_user.password = generate_password_hash(new_pass)
            db.session.commit()
            flash('Password berhasil diubah!', 'success')
            return redirect('/guru/dashboard')
    return render_template('guru/ganti_password.html')